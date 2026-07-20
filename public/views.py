import csv
import json
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from indicateurs.models import Indicateur, DonneeCollectee, MetadonneeIndicateur


def accueil(request):
    return render(request, 'public/accueil.html')


def autocomplete_indicateurs(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse([], safe=False)

    resultats = Indicateur.objects.filter(
        actif=True
    ).filter(
        Q(nom__unaccent__icontains=q) |
        Q(theme__unaccent__icontains=q) |
        Q(sujet__unaccent__icontains=q)
    ).values('id', 'nom', 'source', 'theme')[:8]

    data = [
        {
            'id': r['id'],
            'nom': r['nom'],
            'source': r['source'] or '',
            'theme': r['theme'] or '',
        }
        for r in resultats
    ]
    return JsonResponse(data, safe=False)


def liste_indicateurs(request):
    indicateurs = Indicateur.objects.filter(actif=True).select_related('metadonnee')

    source    = request.GET.get('source', '')
    zone      = request.GET.get('zone', '')
    etage     = request.GET.get('etage', '')
    recherche = request.GET.get('q', '')

    if source:
        indicateurs = indicateurs.filter(source=source)
    if zone:
        if source == 'BASICSET':
            indicateurs = indicateurs.filter(composante_basicset=zone)
        else:
            indicateurs = indicateurs.filter(
                Q(zone_cisat=zone) | Q(composante_basicset=zone)
            )
    if etage:
        indicateurs = indicateurs.filter(etage=etage)
    if recherche:
        indicateurs = indicateurs.filter(
            Q(nom__unaccent__icontains=recherche) |
            Q(theme__unaccent__icontains=recherche) |
            Q(sujet__unaccent__icontains=recherche)
        )

    indicateurs = indicateurs.order_by('nom')

    paginator = Paginator(indicateurs, 20)
    page      = request.GET.get('page')
    page_obj  = paginator.get_page(page)

    zones_cisat = list(Indicateur.objects.filter(source='CISAT', actif=True)
                       .values_list('zone_cisat', flat=True)
                       .distinct().order_by('zone_cisat'))

    composantes_bs = list(Indicateur.objects.filter(source='BASICSET', actif=True)
                          .values_list('composante_basicset', flat=True)
                          .distinct().order_by('composante_basicset'))

    context = {
        'page_obj':       page_obj,
        'zones':          zones_cisat,
        'composantes':    composantes_bs,
        'zones_json':     mark_safe(json.dumps(zones_cisat)),
        'composantes_json': mark_safe(json.dumps(composantes_bs)),
        'source':         source,
        'zone':           zone,
        'etage':          etage,
        'recherche':      recherche,
        'total':          indicateurs.count(),
    }
    return render(request, 'public/liste_indicateurs.html', context)


def _get_filtered_indicateurs(request):
    """Applique les memes filtres que liste_indicateurs et retourne le queryset."""
    indicateurs = Indicateur.objects.filter(actif=True)
    source    = request.GET.get('source', '')
    zone      = request.GET.get('zone', '')
    etage     = request.GET.get('etage', '')
    recherche = request.GET.get('q', '')
    if source:
        indicateurs = indicateurs.filter(source=source)
    if zone:
        if source == 'BASICSET':
            indicateurs = indicateurs.filter(composante_basicset=zone)
        else:
            indicateurs = indicateurs.filter(
                Q(zone_cisat=zone) | Q(composante_basicset=zone)
            )
    if etage:
        indicateurs = indicateurs.filter(etage=etage)
    if recherche:
        indicateurs = indicateurs.filter(
            Q(nom__unaccent__icontains=recherche) |
            Q(theme__unaccent__icontains=recherche) |
            Q(sujet__unaccent__icontains=recherche)
        )
    return indicateurs.order_by('nom')


def _excel_header_style():
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }


def export_indicateurs(request):
    indicateurs = _get_filtered_indicateurs(request)
    fmt = request.GET.get('format', 'csv')

    headers = ['Indicateur', 'Code', 'Nom', 'Theme', 'Zone / Composante', 'Niveau', 'Unite de mesure']
    rows = []
    for ind in indicateurs:
        zone = ind.zone_cisat or (ind.composante_basicset if ind.composante_basicset else '')
        rows.append([
            ind.source,
            ind.code or '',
            ind.nom,
            ind.theme or '',
            zone,
            ind.get_etage_display(),
            ind.unite_mesure or '',
        ])

    if fmt == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Indicateurs'
        style = _excel_header_style()
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="indicateurs_cisat.xlsx"'
        wb.save(response)
        return response

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="indicateurs_cisat.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def export_donnees_indicateur(request, pk):
    indicateur = get_object_or_404(Indicateur, pk=pk, actif=True)
    donnees = DonneeCollectee.objects.filter(
        indicateur=indicateur, statut='valide'
    ).order_by('annee_reference')
    fmt = request.GET.get('format', 'csv')

    headers = ['Indicateur', 'Annee', 'Valeur numerique', 'Valeur texte', 'Unite', 'Source', 'Methode']
    rows = []
    for d in donnees:
        rows.append([
            indicateur.nom,
            d.annee_reference,
            str(d.valeur_numerique) if d.valeur_numerique is not None else '',
            d.valeur_texte or '',
            d.unite_saisie or indicateur.unite_mesure or '',
            d.source_donnee or '',
            d.methode_collecte or '',
        ])

    if fmt == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Donnees'
        style = _excel_header_style()
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        safe_name = indicateur.nom[:30].replace(' ', '_')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="donnees_{safe_name}.xlsx"'
        wb.save(response)
        return response

    safe_name = indicateur.nom[:30].replace(' ', '_')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="donnees_{safe_name}.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def detail_indicateur(request, pk):
    indicateur = get_object_or_404(Indicateur, pk=pk, actif=True)

    try:
        meta = indicateur.metadonnee
    except MetadonneeIndicateur.DoesNotExist:
        meta = None

    donnees = DonneeCollectee.objects.filter(
        indicateur=indicateur,
        statut='valide'
    ).order_by('annee_reference')

    donnees_list = []
    graph_data = []
    annees_disponibles = []
    for d in donnees:
        entry = {
            'annee': d.annee_reference,
            'valeur_num': float(d.valeur_numerique) if d.valeur_numerique is not None else None,
            'valeur_texte': d.valeur_texte or '',
            'unite': d.unite_saisie or indicateur.unite_mesure or '',
            'source': d.source_donnee or '',
        }
        donnees_list.append(entry)
        if d.annee_reference not in annees_disponibles:
            annees_disponibles.append(d.annee_reference)
        if entry['valeur_num'] is not None:
            graph_data.append({'x': d.annee_reference, 'y': entry['valeur_num']})

    context = {
        'indicateur':          indicateur,
        'meta':                meta,
        'donnees':             donnees,
        'donnees_json':        mark_safe(json.dumps(donnees_list)),
        'graph_json':          mark_safe(json.dumps(graph_data)),
        'annees_disponibles':  annees_disponibles,
        'annees_disponibles_json': mark_safe(json.dumps(annees_disponibles)),
        'unite_label':         indicateur.unite_mesure or 'Valeur',
    }
    return render(request, 'public/detail_indicateur.html', context)
